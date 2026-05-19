# -*- coding: utf-8 -*-
"""
Сверка выгрузки «Планирование закупок» (03) с листом «Проставление планов Gremlin»
книги Владимир: воспроизведение логики procurement_planning.gs (якорь месяца,
колонка «Заказали, шт», сумма по строкам с nmId).

Запуск (из папки репозитория или с полными путями к xlsx):
  py -3 tools_reconcile_planning_vs_vladimir.py

Выход:
  reports/plan_reconciliation_report.md
  reports/plan_reconciliation_diffs.csv
"""
from __future__ import annotations

import csv
import math
import re
import sys
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# --- пути по умолчанию (при необходимости замените) ---
P03 = Path(r"c:\Users\Сан Саныч\Downloads\03_Планирование_Закупок (1).xlsx")
PV = Path(r"c:\Users\Сан Саныч\Downloads\Владимир (2026) ВБ - План_факт (1).xlsx")
GREMLIN_SUBSTR = "Gremlin"
TESFE_MARKERS = ("Tesfe", "Тесфе", "Общий")

PP_SCAN_TOP_ROWS = 10
PP_HEADER_SCAN_ROWS = 8
PP_MAX_COLS_FROM_ANCHOR = 45
PP_MONTH_ANCHOR_MAX_PICK_ROW = 4

REPO = Path(__file__).resolve().parent
OUT_DIR = REPO / "reports"


def norm_cell(s):
    if s is None:
        return ""
    return str(s).strip().replace("\u00a0", " ")


def canon(s):
    return re.sub(r"\s+", "", norm_cell(s).lower().replace("ё", "е"))


def is_ordered_header(h):
    c = canon(h)
    if "заказали" not in c or "шт" not in c:
        return False
    if "вдень" in c:
        return False
    return True


def is_wb_header(h):
    c = canon(h)
    if "артикул" in c and ("вб" in c or "wb" in c):
        return True
    if c in ("нм", "nm", "нмид", "nmid"):
        return True
    return False


def parse_first_of_month(val):
    if val is None or val == "":
        return None
    if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
        try:
            if val.day == 1 and 1 <= val.month <= 12:
                return val.year, val.month
        except Exception:
            pass
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            base = datetime(1899, 12, 30)
            d = base + timedelta(days=float(val))
            if d.day == 1 and 1 <= d.month <= 12:
                return d.year, d.month
        except (OverflowError, ValueError):
            pass
    s = norm_cell(val)
    m = re.match(r"^(\d{1,2})[\.\-\/](\d{1,2})[\.\-\/](\d{2,4})", s)
    if not m:
        return None
    d, mo, y = int(m[1]), int(m[2]), int(m[3])
    if y < 100:
        y += 2000
    if d == 1 and 1 <= mo <= 12:
        return y, mo
    return None


def norm_wb_digits(s):
    t = norm_cell(s).replace(" ", "").replace("'", "")
    if not t:
        return ""
    if re.fullmatch(r"\d+", t):
        return t
    if re.fullmatch(r"\d+\.\d*", t):
        try:
            n = float(t)
            r = int(round(n))
            if abs(n - r) > 1e-6:
                return t
            if 1e5 <= r <= 1e16:
                return str(r)
        except ValueError:
            pass
    return t


def canon_wb_key(raw):
    return canon(norm_wb_digits(raw))


def looks_nm_id(raw):
    t = norm_wb_digits(raw)
    return bool(re.fullmatch(r"\d{6,12}", t))


def parse_qty(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v) if math.isfinite(float(v)) else 0.0
    s = norm_cell(v).replace(" ", "").replace(",", ".")
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def round_up_carton(qty, pcs_per_box):
    if qty is None or qty <= 0:
        return 0.0
    if not (pcs_per_box and pcs_per_box > 0):
        return float(qty)
    return math.ceil(qty / pcs_per_box) * pcs_per_box


def ym_key(y, m):
    return f"{y}-{m:02d}"


def read_header_matrix(ws, n_rows, max_col):
    m = []
    for r in range(1, n_rows + 1):
        row = []
        for c in range(1, max_col + 1):
            row.append(ws.cell(r, c).value)
        m.append(row)
    return m


def find_all_month_anchor_meta(ws, year, month, top_rows, max_col):
    meta = []
    seen = set()
    mc = min(max_col, ws.max_column or max_col)
    for c in range(1, mc + 1):
        for r in range(1, min(top_rows, ws.max_row) + 1):
            p = parse_first_of_month(ws.cell(r, c).value)
            if p and p[0] == year and p[1] == month:
                if c - 1 not in seen:
                    seen.add(c - 1)
                    meta.append({"col0": c - 1, "row1": r})
                break
    return meta


def find_wb_and_plan_cols(matrix, anchor_col0):
    n_rows = len(matrix)
    n_cols = len(matrix[0]) if matrix else 0
    c0 = max(0, anchor_col0 - 20)
    c1 = min(n_cols - 1, anchor_col0 + PP_MAX_COLS_FROM_ANCHOR)
    wb_col, header_row = -1, -1
    for r in range(n_rows):
        for c in range(c0, c1 + 1):
            if is_wb_header(matrix[r][c]):
                wb_col, header_row = c, r
                break
        if wb_col >= 0:
            break
    if wb_col < 0:
        for r in range(n_rows):
            for c in range(0, min(anchor_col0, n_cols)):
                if is_wb_header(matrix[r][c]):
                    wb_col, header_row = c, r
                    break
            if wb_col >= 0:
                break
    fallback_hr = min(n_rows - 1, 3)
    scan_row = header_row if header_row >= 0 else fallback_hr
    if wb_col >= 0:
        plan_cap = min(n_cols - 1, max(wb_col + PP_MAX_COLS_FROM_ANCHOR, anchor_col0 + PP_MAX_COLS_FROM_ANCHOR))
    else:
        plan_cap = min(n_cols - 1, anchor_col0 + PP_MAX_COLS_FROM_ANCHOR)
    plan_start = (wb_col + 1) if wb_col >= 0 else anchor_col0
    c_end = plan_cap
    found = []

    def push_col(cc):
        if cc not in found:
            found.append(cc)

    if 0 <= scan_row < n_rows:
        for c in range(plan_start, c_end + 1):
            if is_ordered_header(matrix[scan_row][c]):
                push_col(c)
    if not found and wb_col >= 0:
        for r in range(n_rows):
            if r == scan_row:
                continue
            for c in range(wb_col + 1, c_end + 1):
                if is_ordered_header(matrix[r][c]):
                    push_col(c)
    if not found and wb_col < 0:
        for r in range(n_rows):
            if r == scan_row:
                continue
            for c in range(anchor_col0, c_end + 1):
                if is_ordered_header(matrix[r][c]):
                    push_col(c)
    found.sort()
    plan_col = pick_plan_near_anchor(found, anchor_col0)
    hr = header_row if header_row >= 0 else fallback_hr
    return {"wb_col": wb_col, "plan_col": plan_col, "header_row": hr}


def pick_plan_near_anchor(sorted_cols, anchor_col0):
    if not sorted_cols:
        return -1
    ge = [c for c in sorted_cols if c >= anchor_col0]
    if ge:
        return min(ge)
    best = sorted_cols[0]
    best_d = abs(best - anchor_col0)
    for c in sorted_cols[1:]:
        d = abs(c - anchor_col0)
        if d < best_d or (d == best_d and c < best):
            best, best_d = c, d
    return best


def score_anchor(ws, matrix, anchor_col0):
    found = find_wb_and_plan_cols(matrix, anchor_col0)
    if found["plan_col"] < 0 or found["wb_col"] < 0:
        return -1.0
    wb0, pc0 = found["wb_col"], found["plan_col"]
    data_start_row = found["header_row"] + 2
    s = 0.0
    for r in range(data_start_row, min(ws.max_row, data_start_row + 500) + 1):
        raw = ws.cell(r, wb0 + 1).value
        if not looks_nm_id(raw):
            continue
        s += parse_qty(ws.cell(r, pc0 + 1).value)
    return s


def pick_best_anchor_col(ws, year, month, use_row_filter: bool):
    top_rows = min(PP_SCAN_TOP_ROWS, ws.max_row)
    max_col = min(600, ws.max_column or 600)
    meta = find_all_month_anchor_meta(ws, year, month, top_rows, max_col)
    if not meta:
        return None, meta
    if use_row_filter:
        filt = [m for m in meta if m["row1"] <= PP_MONTH_ANCHOR_MAX_PICK_ROW]
        use = filt if filt else meta
    else:
        use = meta
    matrix = read_header_matrix(ws, PP_HEADER_SCAN_ROWS, max_col)
    best_c, best_s = -1, -1.0
    for m in use:
        c0 = m["col0"]
        sc = score_anchor(ws, matrix, c0)
        if sc < 0:
            continue
        if best_c < 0 or sc > best_s or (sc == best_s and c0 < best_c):
            best_c, best_s = c0, sc
    if best_c < 0:
        return (use[0]["col0"] if use else None), meta
    return best_c, meta


def extract_month(ws, year, month, use_row_filter: bool):
    top_rows = min(PP_SCAN_TOP_ROWS, ws.max_row)
    max_col = min(600, ws.max_column or 600)
    anchor0, _ = pick_best_anchor_col(ws, year, month, use_row_filter)
    if anchor0 is None:
        return {}, -1, -1, -1
    matrix = read_header_matrix(ws, PP_HEADER_SCAN_ROWS, max_col)
    found = find_wb_and_plan_cols(matrix, anchor0)
    if found["wb_col"] < 0 or found["plan_col"] < 0:
        return {}, anchor0, found["wb_col"], found["plan_col"]
    wb0, pc0 = found["wb_col"], found["plan_col"]
    data_start = found["header_row"] + 2
    by_art = {}
    for r in range(data_start, ws.max_row + 1):
        raw = ws.cell(r, wb0 + 1).value
        if not norm_cell(str(raw)):
            continue
        k = canon_wb_key(raw)
        if not k:
            continue
        by_art[k] = by_art.get(k, 0.0) + parse_qty(ws.cell(r, pc0 + 1).value)
    return by_art, anchor0, wb0, pc0


def month_from_header_cell(v):
    if hasattr(v, "year") and hasattr(v, "month"):
        return ym_key(v.year, v.month)
    s = norm_cell(v)
    m = re.match(r"^(\d{4})-(\d{2})", s)
    if m:
        return f"{m[1]}-{m[2]}"
    p = parse_first_of_month(v)
    if p:
        return ym_key(p[0], p[1])
    return None


def load_03_planning(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = None
    for n in wb.sheetnames:
        if "ланирование закупок" in n and "расчёт" not in n.lower() and "расчет" not in n.lower():
            sh = wb[n]
            break
    if sh is None:
        wb.close()
        raise RuntimeError("Не найден лист планирования закупок в 03")
    header_row = None
    for r in range(1, 25):
        v1 = norm_cell(sh.cell(r, 1).value)
        if "ртикул" in v1 and "ВБ" in v1 and len(v1) <= 40:
            header_row = r
            break
    if header_row is None:
        wb.close()
        raise RuntimeError("Не найдена строка шапки с Артикул ВБ")
    month_cols = {}
    for c in range(1, sh.max_column + 1):
        mk = month_from_header_cell(sh.cell(header_row, c).value)
        if mk:
            month_cols[mk] = c
    rows = []
    for r in range(header_row + 1, sh.max_row + 1):
        raw_wb = sh.cell(r, 1).value
        k = canon_wb_key(raw_wb)
        if not k or not looks_nm_id(raw_wb):
            continue
        src = norm_cell(sh.cell(r, 5).value)
        per = {}
        for mk, cc in month_cols.items():
            per[mk] = parse_qty(sh.cell(r, cc).value)
        rows.append({"key": k, "display": norm_wb_digits(raw_wb) or norm_cell(str(raw_wb)), "source": src, "months": per})
    wb.close()
    return rows, sorted(month_cols.keys()), month_cols, header_row


def is_gremlin_only_source(src: str) -> bool:
    s = norm_cell(src)
    if not s:
        return False
    if GREMLIN_SUBSTR not in s:
        return False
    for t in TESFE_MARKERS:
        if t in s:
            return False
    if "Ozon" in s or "озон" in s.lower():
        return False
    return True


def is_tesfe_only_source(src: str) -> bool:
    """Только Tesfe/Общий в источнике 03, без Gremlin и без Ozon."""
    s = norm_cell(src)
    if not s:
        return False
    if GREMLIN_SUBSTR in s:
        return False
    if "Ozon" in s or "озон" in s.lower():
        return False
    return "Tesfe" in s or "Тесфе" in s or "Общий" in s


def analyze_sheet(ws, month_keys, filtered_rows, sheet_label: str):
    """Возвращает словарь с картами месяцев, якорями и списками расхождений."""
    old_maps, new_maps, anchor_info = {}, {}, {}
    max_scan = min(600, ws.max_column or 600)
    for mk in month_keys:
        y, m = map(int, mk.split("-"))
        old_maps[mk], a_old, _, _ = extract_month(ws, y, m, use_row_filter=False)
        new_maps[mk], a_new, _, _ = extract_month(ws, y, m, use_row_filter=True)
        meta = find_all_month_anchor_meta(ws, y, m, PP_SCAN_TOP_ROWS, max_scan)
        anchor_info[mk] = {
            "anchors": len(meta),
            "pick_old_col1": a_old + 1 if a_old is not None and a_old >= 0 else None,
            "pick_new_col1": a_new + 1 if a_new is not None and a_new >= 0 else None,
        }
    diffs_old_new = []
    diffs_03_round = []
    PCS = 24
    for r in filtered_rows:
        k = r["key"]
        for mk in month_keys:
            v03 = r["months"].get(mk, 0.0)
            vo = old_maps.get(mk, {}).get(k, 0.0)
            vn = new_maps.get(mk, {}).get(k, 0.0)
            rn = round_up_carton(vn, PCS)
            if abs(vo - vn) > 1e-6:
                diffs_old_new.append(
                    {
                        "sheet": sheet_label,
                        "nm_key": k,
                        "display": r["display"],
                        "source": r["source"],
                        "month": mk,
                        "vladimir_old_raw": vo,
                        "vladimir_new_raw": vn,
                        "delta_raw": vn - vo,
                        "planning_03": v03,
                        "rounded_new_24": rn,
                        "diff_03_minus_rounded": v03 - rn,
                    }
                )
            if (v03 > 0 or vn > 0) and abs(v03 - rn) > 0.5:
                diffs_03_round.append(
                    {
                        "sheet": sheet_label,
                        "nm_key": k,
                        "display": r["display"],
                        "source": r["source"],
                        "month": mk,
                        "planning_03": v03,
                        "vladimir_new_raw": vn,
                        "rounded_new_24": rn,
                        "diff_03_minus_rounded": v03 - rn,
                    }
                )
    return {
        "old_maps": old_maps,
        "new_maps": new_maps,
        "anchor_info": anchor_info,
        "diffs_old_new": diffs_old_new,
        "diffs_03_round": diffs_03_round,
        "PCS": PCS,
    }


def main():
    if not P03.exists() or not PV.exists():
        print("Файлы не найдены. Укажите P03 и PV в tools_reconcile_planning_vs_vladimir.py", file=sys.stderr)
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wbv = openpyxl.load_workbook(PV, data_only=True)
    gname = next((n for n in wbv.sheetnames if GREMLIN_SUBSTR in n), None)
    if not gname:
        wbv.close()
        raise RuntimeError("Лист Gremlin не найден")
    gsh = wbv[gname]

    rows03, month_keys, _, header_row = load_03_planning(P03)

    gremlin_rows = [r for r in rows03 if is_gremlin_only_source(r["source"])]
    tesfe_rows = [r for r in rows03 if is_tesfe_only_source(r["source"])]
    other_rows = len(rows03) - len(gremlin_rows) - len(tesfe_rows)

    ag = analyze_sheet(gsh, month_keys, gremlin_rows, "Gremlin")

    tname = next((n for n in wbv.sheetnames if "Общий" in n and GREMLIN_SUBSTR not in n), None)
    at = analyze_sheet(wbv[tname], month_keys, tesfe_rows, "Tesfe") if tname else None

    diff_all_old = ag["diffs_old_new"] + (at["diffs_old_new"] if at else [])
    diff_all_03 = ag["diffs_03_round"] + (at["diffs_03_round"] if at else [])

    csv_path = OUT_DIR / "plan_reconciliation_diffs.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "kind",
                "sheet",
                "nm_key",
                "display",
                "source",
                "month",
                "vladimir_old_raw",
                "vladimir_new_raw",
                "delta_raw",
                "planning_03",
                "rounded_new_24",
                "diff_03_minus_rounded",
            ],
        )
        w.writeheader()
        for d in diff_all_old:
            w.writerow({**d, "kind": "old_raw_ne_new_raw"})
        for d in diff_all_03:
            w.writerow(
                {
                    "kind": "03_ne_round_new_24",
                    "sheet": d["sheet"],
                    "nm_key": d["nm_key"],
                    "display": d["display"],
                    "source": d["source"],
                    "month": d["month"],
                    "vladimir_old_raw": "",
                    "vladimir_new_raw": d["vladimir_new_raw"],
                    "delta_raw": "",
                    "planning_03": d["planning_03"],
                    "rounded_new_24": d["rounded_new_24"],
                    "diff_03_minus_rounded": d["diff_03_minus_rounded"],
                }
            )

    PCS = ag["PCS"]
    n_g = len(gremlin_rows)
    n_t = len(tesfe_rows)
    cells_g = n_g * len(month_keys)
    cells_t = n_t * len(month_keys) if at else 0
    by_m_old_new_g = Counter(d["month"] for d in ag["diffs_old_new"])
    by_m_03_g = Counter(d["month"] for d in ag["diffs_03_round"])
    n_old_new_g = len({(d["nm_key"], d["month"]) for d in ag["diffs_old_new"]})
    n_03_mis_g = len({(d["nm_key"], d["month"]) for d in ag["diffs_03_round"]})
    if at:
        by_m_old_new_t = Counter(d["month"] for d in at["diffs_old_new"])
        by_m_03_t = Counter(d["month"] for d in at["diffs_03_round"])
        n_old_new_t = len({(d["nm_key"], d["month"]) for d in at["diffs_old_new"]})
        n_03_mis_t = len({(d["nm_key"], d["month"]) for d in at["diffs_03_round"]})
    else:
        by_m_old_new_t = Counter()
        by_m_03_t = Counter()
        n_old_new_t = 0
        n_03_mis_t = 0

    md = []
    md.append("# Сверка: 03 «Планирование закупок» vs Владимир (Gremlin + Tesfe)\n")
    md.append("## Исходные файлы\n")
    md.append(f"- **03:** `{P03}`\n")
    md.append(f"- **Владимир:** `{PV}`\n")
    md.append(f"- **Лист Gremlin:** `{gname}`\n")
    if tname:
        md.append(f"- **Лист Tesfe (Общий):** `{tname}`\n")
    else:
        md.append("- **Лист Tesfe:** не найден (имя без «Gremlin», с «Общий»).\n")
    md.append(f"- **Строка шапки артикулов в 03:** {header_row}\n")
    md.append(f"- **Месяцы в шапке 03:** {', '.join(month_keys)}\n")
    md.append("\n## Методика\n")
    md.append(
        "1. Для каждого листа Владимир и месяца горизонта воспроизведена логика `procurement_planning.gs`: якорь 01.MM, "
        "сумма «Заказали, шт» по nmId, колонки «Артикул ВБ» и «Заказали, шт» (в т.ч. не левее якоря). "
        "Для «Общий»: если «Артикул ВБ» вне полосы ±20 от якоря, выполняется **поиск ВБ слева от якоря** и расширенная полоса до `max(wb+45, anchor+45)` — как в обновлённом Apps Script.\n"
    )
    md.append(
        f"2. **«Новый» якорь** — даты 01.MM в строке **> {PP_MONTH_ANCHOR_MAX_PICK_ROW}** отбрасываются при выборе, если есть кандидаты выше.\n"
    )
    md.append("3. **«Старый» якорь** — все даты месяца участвуют в максимуме суммы.\n")
    md.append(
        "4. Строки **03**: для Gremlin — только источник Gremlin; для Tesfe — только Tesfe/Общий без Gremlin и без Ozon.\n"
    )
    md.append(
        f"5. Округление для сравнения с 03: `ceil(сырой / {PCS}) * {PCS}` (упрощение; в скрипте — коробка из справочника).\n"
    )

    def append_block(title, anchor_info, diffs_old_new, diffs_03_round, n_rows, cells_n, n_old_new, n_03_mis, by_m_old_new, by_m_03):
        lines = [f"\n## {title}\n"]
        lines.append("### Краткие итоги\n")
        lines.append("| Показатель | Значение |\n|---|---|\n")
        lines.append(f"| Строк 03 в этой сверке | {n_rows} |\n")
        lines.append(f"| Ячеек артикул×месяц | {cells_n} |\n")
        lines.append(f"| old_raw ≠ new_raw | {n_old_new} |\n")
        lines.append(f"| 03 ≠ округление(new, {PCS}) при |Δ|>0.5 | {n_03_mis} |\n")
        lines.append("\n### old vs new по месяцам\n")
        lines.append("| Месяц | old ≠ new |\n|---|---:|\n")
        for mk in month_keys:
            lines.append(f"| {mk} | {by_m_old_new.get(mk, 0)} |\n")
        lines.append("\n### 03 vs округление по месяцам\n")
        lines.append("| Месяц | |Δ|>0.5 |\n|---|---:|\n")
        for mk in month_keys:
            lines.append(f"| {mk} | {by_m_03.get(mk, 0)} |\n")
        lines.append("\n### Якорь (колонка 1-based)\n")
        lines.append("| Месяц | Число дат 01.MM | Старый col | Новый col |\n|---|---|---|---|\n")
        for mk in month_keys:
            ai = anchor_info[mk]
            lines.append(f"| {mk} | {ai['anchors']} | {ai['pick_old_col1']} | {ai['pick_new_col1']} |\n")
        lines.append("\n### Примеры old ≠ new (до 25)\n")
        lines.append("| Артикул | Месяц | Было | Стало | Δ | 03 | Округл. |\n|---:|---:|---:|---:|---:|---:|---:|\n")
        for d in diffs_old_new[:25]:
            lines.append(
                f"| {d['display']} | {d['month']} | {d['vladimir_old_raw']:.0f} | {d['vladimir_new_raw']:.0f} | "
                f"{d['delta_raw']:.0f} | {d['planning_03']:.0f} | {d['rounded_new_24']:.0f} |\n"
            )
        md.extend(lines)

    md.append("\n## Сводка по файлу 03\n")
    md.append(f"| Показатель | Значение |\n|---|---|\n")
    md.append(f"| Строк с nmId | {len(rows03)} |\n")
    md.append(f"| Только Gremlin | {n_g} |\n")
    md.append(f"| Только Tesfe | {n_t} |\n")
    md.append(f"| Прочие / смешанные источники | {other_rows} |\n")

    append_block(
        "Gremlin («Проставление планов Gremlin»)",
        ag["anchor_info"],
        ag["diffs_old_new"],
        ag["diffs_03_round"],
        n_g,
        cells_g,
        n_old_new_g,
        n_03_mis_g,
        by_m_old_new_g,
        by_m_03_g,
    )
    if at:
        append_block(
            f"Tesfe («{tname}»)",
            at["anchor_info"],
            at["diffs_old_new"],
            at["diffs_03_round"],
            n_t,
            cells_t,
            n_old_new_t,
            n_03_mis_t,
            by_m_old_new_t,
            by_m_03_t,
        )
    else:
        md.append("\n## Tesfe\nЛист «Проставление планов Общий» не найден — блок пропущен.\n")

    md.append("\n## Выявленные проблемы (до правки кода)\n")
    md.append(
        "- На листе Gremlin для одного и того же месяца встречается **несколько** ячеек с датой 1-го числа в **разных колонках** "
        "и иногда в **разных строках** шапки.\n"
    )
    md.append(
        "- Алгоритм «максимум суммы по nmId» выбирал **нижний** (например, строка 5) дубль месяца с **другой** колонкой «Заказали, шт», "
        "из-за чего по многим артикулам подставлялись завышенные числа; после округления до коробки расхождение с визуально «правильным» блоком усиливалось.\n"
    )
    md.append(
        "- На листе **Общий (Tesfe)** даты месяца часто **далеко справа**, а «Артикул ВБ» — **слева**; без поиска ВБ слева от якоря план с листа **не читался** (в xlsx ранее `WB_col=-1`). В `procurement_planning.gs` добавлен такой fallback и расширена полоса «Заказали, шт» до `max(wb+45, anchor+45)`.\n"
    )
    md.append("\n## Ограничения сверки\n")
    md.append(
        "- Сравнение с **03** через фиксированные **24 шт/кор** — упрощение; в `procurement_planning.gs` коробка из справочника, часть расхождений — **только из-за коробки**.\n"
    )
    md.append(
        "- Выгрузка **xlsx** может отличаться от текущих Google Таблиц.\n"
    )
    md.append(
        "- Строки **Gremlin+Tesfe** в отчёте не сверяются как сумма двух листов (нужна отдельная логика).\n"
    )

    md.append(f"\nПолный список отличий — в CSV (колонка **sheet**): `{csv_path.relative_to(REPO)}`.\n")
    md.append("\n## Проделанная работа\n")
    md.append(
        "- Реализован автономный расчёт на Python по структуре xlsx (openpyxl), без доступа к Google API.\n"
        "- Воспроизведены: поиск якорей, оценка суммы, выбор колонок ВБ/«Заказали, шт», суммирование по строкам с nmId.\n"
        "- Добавлена сверка листа **Tesfe (Общий)** и колонка **sheet** в CSV.\n"
    )

    report_path = OUT_DIR / "plan_reconciliation_report.md"
    report_path.write_text("".join(md), encoding="utf-8")

    wbv.close()
    print("OK:", report_path)
    print("CSV:", csv_path)


if __name__ == "__main__":
    main()
