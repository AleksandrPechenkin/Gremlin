# -*- coding: utf-8 -*-
import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

P = Path(r"c:\Users\Сан Саныч\Downloads\Владимир Озон - План_факт.xlsx")


def norm(s):
    return "" if s is None else str(s).strip()


def canon(s):
    return re.sub(r"\s+", "", norm(s).lower().replace("ё", "е"))


def is_ordered(h):
    c = canon(h)
    return "заказали" in c and "шт" in c and "вдень" not in c


def is_supplier(h):
    c = canon(h)
    return "артикул" in c and "поставщ" in c and "вб" not in c and "wb" not in c


def parse_month(v):
    if hasattr(v, "year") and hasattr(v, "day"):
        if v.day == 1:
            return v.year, v.month
    if isinstance(v, (int, float)):
        d = datetime(1899, 12, 30) + timedelta(days=float(v))
        if d.day == 1:
            return d.year, d.month
    s = norm(v)
    m = re.match(r"^(\d{1,2})[\.\-/](\d{1,2})[\.\-/](\d{2,4})", s)
    if m:
        d, mo, y = int(m[1]), int(m[2]), int(m[3])
        if y < 100:
            y += 2000
        if d == 1:
            return y, mo
    return None


wb = openpyxl.load_workbook(P, data_only=True)
sh = wb["Проставление планов"]
print("dims", sh.max_row, sh.max_column)
for r in range(1, 25):
    cols = []
    for c in range(1, 120):
        v = sh.cell(r, c).value
        if is_supplier(v) or is_ordered(v) or parse_month(v):
            cols.append((c, repr(v)[:50]))
    if cols:
        print("row", r, cols[:25])
for r in range(10, 40):
    v = sh.cell(r, 15).value
    if v and norm(v) and "ртикул" not in norm(v):
        print("sample col15 r", r, repr(v))
        break
wb.close()
