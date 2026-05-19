# -*- coding: utf-8 -*-
"""
Сверка планов Ozon: лист «03 Планирование закупок» (источник Ozon) vs
«Проставление планов» в «Владимир Озон - План_факт.xlsx».

Раскладка Ozon — «Артикул поставщика» + «Заказали, шт» (как ppExtractPlansSupplierLayoutBatch_).

Запуск:
  py -3 tools_reconcile_ozon_planning.py

Выход:
  reports/plan_reconciliation_ozon_report.md
  reports/plan_reconciliation_ozon_diffs.csv
"""
from __future__ import annotations

import csv
import math
import re
import sys
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

P03 = Path(r"c:\Users\Сан Саныч\Downloads\03_Планирование_Закупок (1).xlsx")
POZON = Path(r"c:\Users\Сан Саныч\Downloads\Владимир Озон - План_факт.xlsx")
OZON_SHEET = "Проставление планов"

PP_SCAN_TOP_ROWS = 10
PP_HEADER_SCAN_ROWS = 8
PP_MAX_COLS_FROM_ANCHOR = 45
PP_MONTH_ANCHOR_MAX_PICK_ROW = 4
PP_ANCHOR_SCORE_MAX_BODY_ROWS = 250

REPO = Path(__file__).resolve().parent
OUT_DIR = REPO / "reports"


def norm_cell(s):
    if s is None:
        return ""
    return str(s).strip().replace("\u00a0", " ")


def canon(s):
    return re.sub(r"\s+", "", norm_cell(s).lower().replace("ё", "е"))


def is_ordered_header(h):
    c = canon(h)
    return "заказали" in c and "шт" in c and "вдень" not in c


def is_supplier_header(h):
    c = canon(h)
    return "артикул" in c and "поставщ" in c and "вб" not in c and "wb" not in c


def norm_wb_digits(s):
    t = norm_cell(s).replace(" ", "").replace("'", "")
    if not t:
        return ""
    if re.fullmatch(r"\d+", t):
        return t
    if re.fullmatch(r"\d+\.\d*", t):
        try:
            n = float(t)
            r = int(round(n))
            if abs(n - r) <= 1e-6 and 1e5 <= r <= 1e16:
                return str(r)
        except ValueError:
            pass
    return t


def canon_wb_key(raw):
    return canon(norm_wb_digits(raw))


def looks_nm_id(raw):
    return bool(re.fullmatch(r"\d{6,12}", norm_wb_digits(raw)))


def parse_qty(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v) if math.isfinite(float(v)) else 0.0
    s = norm_cell(v).replace(" ", "").replace(",", ".")
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def round_up_carton(qty, pcs):
    if qty is None or qty <= 0:
        return 0.0
    if not (pcs and pcs > 0):
        return float(qty)
    return math.ceil(qty / pcs) * pcs


def ym_key(y, m):
    return f"{y}-{m:02d}"


def parse_first_of_month(val):
    if val is None or val == "":
        return None
    if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
        try:
            if val.day == 1 and 1 <= val.month <= 12:
                return val.year, val.month
        except Exception:
            pass
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            d = datetime(1899, 12, 30) + timedelta(days=float(val))
            if d.day == 1 and 1 <= d.month <= 12:
                return d.year, d.month
        except (OverflowError, ValueError):
            pass
    s = norm_cell(val)
    m = re.match(r"^(\d{1,2})[\.\-\/](\d{1,2})[\.\-\/](\d{2,4})", s)
    if not m:
        return None
    d, mo, y = int(m[1]), int(m[2]), int(m[3])
    if y < 100:
        y += 2000
    if d == 1 and 1 <= mo <= 12:
        return y, mo
    return None


def read_header_matrix(ws, n_rows, max_col):
    m = []
    for r in range(1, n_rows + 1):
        row = []
        for c in range(1, max_col + 1):
            row.append(ws.cell(r, c).value)
        m.append(row)
    return m


def find_all_month_anchor_meta(ws, year, month, top_rows, max_col):
    meta = []
    seen = set()
    mc = min(max_col, ws.max_column or max_col)
    for c in range(1, mc + 1):
        for r in range(1, min(top_rows, ws.max_row) + 1):
            p = parse_first_of_month(ws.cell(r, c).value)
            if p and p[0] == year and p[1] == month:
                if c - 1 not in seen:
                    seen.add(c - 1)
                    meta.append({"col0": c - 1, "row1": r})
                break
    return meta


def pick_plan_near_anchor(sorted_cols, anchor_col0):
    if not sorted_cols:
        return -1
    ge = [c for c in sorted_cols if c >= anchor_col0]
    if ge:
        return min(ge)
    best = sorted_cols[0]
    best_d = abs(best - anchor_col0)
    for c in sorted_cols[1:]:
        d = abs(c - anchor_col0)
        if d < best_d or (d == best_d and c < best):
            best, best_d = c, d
    return best


def find_supplier_sku_and_ordered_cols(matrix, anchor_col0):
    n_rows = len(matrix)
    n_cols = len(matrix[0]) if matrix else 0
    c0 = max(0, anchor_col0 - 20)
    c1 = min(n_cols - 1, anchor_col0 + PP_MAX_COLS_FROM_ANCHOR)
    best = {"sku_col": -1, "plan_col": -1, "header_row": -1}
    for r in range(n_rows):
        sku_col, plan_col = -1, -1
        for c in range(c0, c1 + 1):
            h = matrix[r][c]
            if is_supplier_header(h):
                sku_col = c
            if is_ordered_header(h):
                plan_col = c
        if sku_col >= 0 and plan_col >= 0:
            return {"sku_col": sku_col, "plan_col": plan_col, "header_row": r}
        if sku_col >= 0 and best["sku_col"] < 0:
            best = {"sku_col": sku_col, "plan_col": -1, "header_row": r}
    return best


def score_anchor_supplier(ws, matrix, anchor_col0, supplier_to_wb):
    found = find_supplier_sku_and_ordered_cols(matrix, anchor_col0)
    if found["plan_col"] < 0 or found["sku_col"] < 0:
        return -1.0
    sku0, pc0 = found["sku_col"], found["plan_col"]
    data_start = found["header_row"] + 2
    s = 0.0
    cap = min(ws.max_row, data_start + PP_ANCHOR_SCORE_MAX_BODY_ROWS)
    for r in range(data_start, cap + 1):
        raw = ws.cell(r, sku0 + 1).value
        sk = canon(norm_cell(raw))
        if not sk or sk not in supplier_to_wb:
            continue
        s += parse_qty(ws.cell(r, pc0 + 1).value)
    return s


def pick_best_anchor_supplier(ws, year, month, supplier_to_wb, use_row_filter):
    top_rows = min(PP_SCAN_TOP_ROWS, ws.max_row)
    max_col = min(600, ws.max_column or 600)
    meta = find_all_month_anchor_meta(ws, year, month, top_rows, max_col)
    if not meta:
        return None, meta
    if use_row_filter:
        filt = [m for m in meta if m["row1"] <= PP_MONTH_ANCHOR_MAX_PICK_ROW]
        use = filt if filt else meta
    else:
        use = meta
    matrix = read_header_matrix(ws, PP_HEADER_SCAN_ROWS, max_col)
    best_c, best_s = -1, -1.0
    for m in use:
        c0 = m["col0"]
        sc = score_anchor_supplier(ws, matrix, c0, supplier_to_wb)
        if sc < 0:
            continue
        if best_c < 0 or sc > best_s or (sc == best_s and c0 < best_c):
            best_c, best_s = c0, sc
    if best_c < 0:
        return (use[0]["col0"] if use else None), meta
    return best_c, meta


def extract_month_supplier(ws, year, month, supplier_to_wb, use_row_filter, by_supplier_key=False):
    top_rows = min(PP_SCAN_TOP_ROWS, ws.max_row)
    max_col = min(600, ws.max_column or 600)
    anchor0, _ = pick_best_anchor_supplier(ws, year, month, supplier_to_wb, use_row_filter)
    if anchor0 is None:
        return {}, -1, -1, -1, 0
    matrix = read_header_matrix(ws, PP_HEADER_SCAN_ROWS, max_col)
    found = find_supplier_sku_and_ordered_cols(matrix, anchor0)
    if found["sku_col"] < 0 or found["plan_col"] < 0:
        return {}, anchor0, found["sku_col"], found["plan_col"], 0
    sku0, pc0 = found["sku_col"], found["plan_col"]
    data_start = found["header_row"] + 2
    by_out = {}
    unmapped = 0
    for r in range(data_start, ws.max_row + 1):
        raw = ws.cell(r, sku0 + 1).value
        sk = canon(norm_cell(raw))
        if not sk:
            continue
        if by_supplier_key:
            out_key = sk
        else:
            out_key = supplier_to_wb.get(sk)
            if not out_key:
                unmapped += 1
                continue
        by_out[out_key] = by_out.get(out_key, 0.0) + parse_qty(ws.cell(r, pc0 + 1).value)
    return by_out, anchor0, sku0, pc0, unmapped


def month_from_header_cell(v):
    if hasattr(v, "year") and hasattr(v, "month"):
        return ym_key(v.year, v.month)
    s = norm_cell(v)
    m = re.match(r"^(\d{4})-(\d{2})", s)
    if m:
        return f"{m[1]}-{m[2]}"
    p = parse_first_of_month(v)
    if p:
        return ym_key(p[0], p[1])
    return None


def load_03_planning(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = None
    for n in wb.sheetnames:
        if "ланирование закупок" in n and "расчёт" not in n.lower() and "расчет" not in n.lower():
            sh = wb[n]
            break
    if sh is None:
        wb.close()
        raise RuntimeError("Не найден лист планирования в 03")
    header_row = None
    for r in range(1, 25):
        v1 = norm_cell(sh.cell(r, 1).value)
        if "ртикул" in v1 and "ВБ" in v1 and len(v1) <= 40:
            header_row = r
            break
    if header_row is None:
        wb.close()
        raise RuntimeError("Шапка 03 не найдена")
    month_cols = {}
    for c in range(1, sh.max_column + 1):
        mk = month_from_header_cell(sh.cell(header_row, c).value)
        if mk:
            month_cols[mk] = c
    rows = []
    for r in range(header_row + 1, sh.max_row + 1):
        raw_wb = sh.cell(r, 1).value
        src = norm_cell(sh.cell(r, 5).value)
        supplier = norm_cell(sh.cell(r, 3).value)
        if not supplier and not norm_cell(str(raw_wb)):
            continue
        if "источник" in src.lower() and "план" in src.lower():
            continue
        per = {mk: parse_qty(sh.cell(r, cc).value) for mk, cc in month_cols.items()}
        rows.append(
            {
                "key": canon_wb_key(raw_wb) if looks_nm_id(raw_wb) else canon(supplier),
                "display": norm_wb_digits(raw_wb) or norm_cell(str(raw_wb)),
                "source": src,
                "supplier": supplier,
                "supplier_key": canon(supplier),
                "months": per,
            }
        )
    wb.close()
    return rows, sorted(month_cols.keys()), header_row


def load_03_ozon_rows(rows03):
    """Строки 03 только с источником Ozon; ключ — артикул поставщика (в 03 кол. A часто OZN…, не nmId)."""
    out = []
    for r in rows03:
        if not is_ozon_only_source(r["source"]):
            continue
        sk = r.get("supplier_key") or canon(r.get("supplier") or "")
        if not sk:
            continue
        out.append(
            {
                "key": sk,
                "display": r.get("supplier") or sk,
                "wb_display": r.get("display") or "",
                "source": r["source"],
                "supplier": r.get("supplier") or "",
                "months": r["months"],
            }
        )
    return out


def is_ozon_only_source(src: str) -> bool:
    s = norm_cell(src)
    if not s:
        return False
    if "ozon" not in s.lower() and "озон" not in s.lower():
        return False
    if "Gremlin" in s or "Tesfe" in s or "Тесфе" in s or "Общий" in s:
        return False
    return True


def build_supplier_to_wb(rows03):
    stw = {}
    for r in rows03:
        sup = canon(r.get("supplier") or "")
        if sup and r["key"]:
            if sup not in stw:
                stw[sup] = r["key"]
    return stw


def main():
    if not P03.exists():
        print("Нет файла 03:", P03, file=sys.stderr)
        sys.exit(1)
    if not POZON.exists():
        print("Нет файла Ozon:", POZON, file=sys.stderr)
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows03, month_keys, header_row = load_03_planning(P03)
    ozon_rows = load_03_ozon_rows(rows03)
    supplier_to_wb = build_supplier_to_wb(rows03)

    wb = openpyxl.load_workbook(POZON, data_only=True)
    if OZON_SHEET not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"Лист «{OZON_SHEET}» не найден")
    osh = wb[OZON_SHEET]

    old_maps, new_maps, anchor_info = {}, {}, {}
    unmapped_by_month = {}
    for mk in month_keys:
        y, m = map(int, mk.split("-"))
        old_maps[mk], a_old, sku_o, pc_o, um_o = extract_month_supplier(
            osh, y, m, supplier_to_wb, use_row_filter=False, by_supplier_key=True
        )
        new_maps[mk], a_new, sku_n, pc_n, um_n = extract_month_supplier(
            osh, y, m, supplier_to_wb, use_row_filter=True, by_supplier_key=True
        )
        unmapped_by_month[mk] = um_n
        meta = find_all_month_anchor_meta(osh, y, m, PP_SCAN_TOP_ROWS, min(600, osh.max_column or 600))
        anchor_info[mk] = {
            "anchors": len(meta),
            "pick_old_col1": a_old + 1 if a_old is not None and a_old >= 0 else None,
            "pick_new_col1": a_new + 1 if a_new is not None and a_new >= 0 else None,
            "sku_col1": sku_n + 1 if sku_n is not None and sku_n >= 0 else None,
            "plan_col1": pc_n + 1 if pc_n is not None and pc_n >= 0 else None,
            "unmapped_rows": um_n,
        }

    PCS = 24
    diffs_old_new = []
    diffs_03 = []
    for r in ozon_rows:
        k = r["key"]  # canon(артикул поставщика)
        for mk in month_keys:
            v03 = r["months"].get(mk, 0.0)
            vo = old_maps.get(mk, {}).get(k, 0.0)
            vn = new_maps.get(mk, {}).get(k, 0.0)
            rn = round_up_carton(vn, PCS)
            if abs(vo - vn) > 1e-6:
                diffs_old_new.append(
                    {
                        "nm_key": k,
                        "display": r["display"],
                        "wb_display": r.get("wb_display", ""),
                        "supplier": r["supplier"],
                        "month": mk,
                        "vladimir_old_raw": vo,
                        "vladimir_new_raw": vn,
                        "delta_raw": vn - vo,
                        "planning_03": v03,
                        "rounded_new_24": rn,
                        "diff_03_minus_rounded": v03 - rn,
                    }
                )
            if (v03 > 0 or vn > 0) and abs(v03 - rn) > 0.5:
                diffs_03.append(
                    {
                        "nm_key": k,
                        "display": r["display"],
                        "wb_display": r.get("wb_display", ""),
                        "supplier": r["supplier"],
                        "month": mk,
                        "planning_03": v03,
                        "vladimir_new_raw": vn,
                        "rounded_new_24": rn,
                        "diff_03_minus_rounded": v03 - rn,
                    }
                )

    csv_path = OUT_DIR / "plan_reconciliation_ozon_diffs.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "kind",
                "nm_key",
                "display",
                "wb_display",
                "supplier",
                "month",
                "vladimir_old_raw",
                "vladimir_new_raw",
                "delta_raw",
                "planning_03",
                "rounded_new_24",
                "diff_03_minus_rounded",
            ],
        )
        w.writeheader()
        for d in diffs_old_new:
            w.writerow({**d, "kind": "old_raw_ne_new_raw"})
        for d in diffs_03:
            w.writerow(
                {
                    "kind": "03_ne_round_new_24",
                    "nm_key": d["nm_key"],
                    "display": d["display"],
                    "wb_display": d.get("wb_display", ""),
                    "supplier": d["supplier"],
                    "month": d["month"],
                    "vladimir_old_raw": "",
                    "vladimir_new_raw": d["vladimir_new_raw"],
                    "delta_raw": "",
                    "planning_03": d["planning_03"],
                    "rounded_new_24": d["rounded_new_24"],
                    "diff_03_minus_rounded": d["diff_03_minus_rounded"],
                }
            )

    n_old_new = len({(d["nm_key"], d["month"]) for d in diffs_old_new})
    n_03 = len({(d["nm_key"], d["month"]) for d in diffs_03})
    by_m_old = Counter(d["month"] for d in diffs_old_new)
    by_m_03 = Counter(d["month"] for d in diffs_03)

    md = []
    md.append("# Сверка Ozon: 03 «Планирование закупок» vs Владимир Озон\n\n")
    md.append(f"- **03:** `{P03}`\n")
    md.append(f"- **Владимир Ozon:** `{POZON}` / лист «{OZON_SHEET}»\n")
    md.append(f"- **Строка шапки 03:** {header_row}\n")
    md.append(f"- **Месяцы:** {', '.join(month_keys)}\n")
    md.append(f"- **Строк 03 с источником Ozon:** {len(ozon_rows)}\n")
    md.append(f"- **Пар supplier→ВБ из 03 (кол. «Наименование»):** {len(supplier_to_wb)}\n\n")
    md.append("## Методика\n\n")
    md.append(
        "1. Раскладка **Ozon** — блоки «Артикул поставщика» + «Заказали, шт» у якоря 01.MM "
        "(как `ppExtractPlansSupplierLayoutBatch_`), не nmId ВБ.\n"
    )
    md.append(
        "2. Сопоставление с 03: артикул поставщика из листа Владимир → ключ ВБ через таблицу из 03 "
        "(колонки «Артикул ВБ» + «Наименование»). В проде — справочник `PRODUCT_REFERENCE`.\n"
    )
    md.append(
        f"3. **Новый якорь** — даты 01.MM только в строках ≤ {PP_MONTH_ANCHOR_MAX_PICK_ROW}, если есть кандидаты выше.\n"
    )
    md.append(f"4. Округление для сравнения с 03: ceil(raw/{PCS})×{PCS} (упрощение; в скрипте — коробка из справочника).\n\n")

    md.append("## Итоги\n\n| Показатель | Значение |\n|---|---|\n")
    md.append(f"| Ячеек артикул×месяц (Ozon в 03) | {len(ozon_rows) * len(month_keys)} |\n")
    md.append(f"| old_raw ≠ new_raw (пары) | {n_old_new} |\n")
    md.append(f"| 03 ≠ округл. new при \\|Δ\\|>0.5 | {n_03} |\n\n")

    md.append("### Якорь и колонки (новый алгоритм)\n\n")
    md.append("| Месяц | #01.MM | Старый col | Новый col | SKU col | Plan col | Строк без маппинга |\n")
    md.append("|---|---:|---:|---:|---:|---:|---:|\n")
    for mk in month_keys:
        ai = anchor_info[mk]
        md.append(
            f"| {mk} | {ai['anchors']} | {ai['pick_old_col1']} | {ai['pick_new_col1']} | "
            f"{ai['sku_col1']} | {ai['plan_col1']} | {ai['unmapped_rows']} |\n"
        )

    md.append("\n### old ≠ new по месяцам\n\n| Месяц | count |\n|---|---:|\n")
    for mk in month_keys:
        md.append(f"| {mk} | {by_m_old.get(mk, 0)} |\n")

    md.append("\n### 03 vs округление по месяцам\n\n| Месяц | count |\n|---|---:|\n")
    for mk in month_keys:
        md.append(f"| {mk} | {by_m_03.get(mk, 0)} |\n")

    md.append("\n### Примеры 03 ≠ округл. new (до 30)\n\n")
    md.append("| Артикул ВБ | Поставщик | Мес | 03 | Сырой Вл | Округл | Δ |\n")
    md.append("|---:|---|---:|---:|---:|---:|---:|\n")
    for d in sorted(diffs_03, key=lambda x: -abs(x["diff_03_minus_rounded"]))[:30]:
        md.append(
            f"| {d['display']} | {d['supplier'][:20]} | {d['month']} | {d['planning_03']:.0f} | "
            f"{d['vladimir_new_raw']:.0f} | {d['rounded_new_24']:.0f} | {d['diff_03_minus_rounded']:.0f} |\n"
        )

    md.append("\n### Примеры смены якоря old→new (до 20)\n\n")
    md.append("| Артикул | Мес | Было | Стало | Δ | 03 |\n|---:|---:|---:|---:|---:|---:|\n")
    for d in diffs_old_new[:20]:
        md.append(
            f"| {d['display']} | {d['month']} | {d['vladimir_old_raw']:.0f} | {d['vladimir_new_raw']:.0f} | "
            f"{d['delta_raw']:.0f} | {d['planning_03']:.0f} |\n"
        )

    md.append("\n## Вероятные причины расхождений\n\n")
    md.append(
        "- На листе **несколько блоков** одного месяца (колонки 8, 63–72, 112+…) — неверный якорь даёт чужой «Заказали, шт».\n"
    )
    md.append(
        "- **Артикул поставщика** в xlsx не совпадает с «Наименованием» в 03/справочнике → строка не попадает в сумму (см. «Строк без маппинга»).\n"
    )
    md.append(
        "- В 03 план **округлён до коробки** из справочника; сверка с сырым Владимиром и фикс. 24 шт завышает ложные «расхождения».\n"
    )
    md.append(f"\nПолный CSV: `{csv_path.relative_to(REPO)}`\n")

    report_path = OUT_DIR / "plan_reconciliation_ozon_report.md"
    report_path.write_text("".join(md), encoding="utf-8")
    wb.close()

    print("OK:", report_path)
    print("CSV:", csv_path)
    print("Ozon rows in 03:", len(ozon_rows))
    print("old!=new pairs:", n_old_new)
    print("03!=round pairs:", n_03)


if __name__ == "__main__":
    main()
